import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def personalizacion_tablas(ruta_xlsx:str):
    # TODO: Cargar el archivo con openpyxl para aplicar estilos
    wb = load_workbook(ruta_xlsx)

    # TODO: Aplicar estilos en cada hoja
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # TODO: Definir estilos
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")      
        
        cell_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell_border = Border(left=Side(style='thick', color="FFFFFF"), 
                            right=Side(style='thick', color="FFFFFF"), 
                            top=Side(style='thick', color="FFFFFF"), 
                            bottom=Side(style='thick', color="FFFFFF"))

        # TODO: Aplicar estilos a los encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = cell_border
            cell.alignment = header_alignment

        # TODO: Aplicar estilos a las celdas del DataFrame
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = cell_fill
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # TODO: Ajustar el ancho de las columnas basado en el contenido
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width

    # TODO Guardar el archivo Excel con estilos
    wb.save(ruta_xlsx)